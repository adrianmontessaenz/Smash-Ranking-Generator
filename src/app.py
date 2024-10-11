# Import libraries
from gql import Client, gql
from gql.transport.requests import RequestsHTTPTransport
import json
import sys
import pandas as pd
import os
from openpyxl import load_workbook

from data_manager import add_tournament, add_head2head, compute_ranking

# Function to fetch tournament data and store it in a JSON file
def fetch_tournament(slug, eventSlug):
    # Set up the GraphQL client
    transport = RequestsHTTPTransport(
        url="https://api.start.gg/gql/alpha",
        headers={"Authorization": "Bearer dc6a93f91afd04e48bb932193e03882e"},  # Replace with your actual API token
        use_json=True,
    )
    client = Client(transport=transport, fetch_schema_from_transport=True)

    # Query with entrants
    query = gql("""
    query getEventData($slug: String, $eventSlug: String) {
      tournament(slug: $slug) {
        events(filter:{slug:$eventSlug})
        {
          id
        	entrants(query:{})
        	{
          	pageInfo {
            	total
          	}
        	}
        }
      }
    }
    """)

    # Get entrants to retrieve all placements
    entrants_data = client.execute(query, variable_values={"slug": slug, "eventSlug": eventSlug})
    total_entrants = entrants_data['tournament']['events'][0]['entrants']['pageInfo']['total']
    

    # Query to get standings
    standings_query = gql("""
    query getStandings($slug: String!, $perPage: Int!, $page: Int! $eventSlug: String!) {
      tournament(slug: $slug) {
        name
        events(filter: {slug: $eventSlug}) {
          entrants(query:{perPage: $perPage, page: $page})  {
            pageInfo {
            	total
          	}
            nodes{
              id
              participants{
                gamerTag
              }
              standing {
                placement
              }
              paginatedSets {
                nodes
                {
                  winnerId
                  slots{
                    entrant{
                      participants{
                        gamerTag
                      }
                    }
                  }                
                }
            	}
        		}
          }
        }
      }
    }
    """)
    
    # Get placements and join jsons
    tmp_counter = 0
    tmp_page_counter = 1
    final_json = []
    while tmp_counter < total_entrants:
      json_data = client.execute(standings_query, variable_values={"slug": slug, "perPage": 10, "page": tmp_page_counter, "eventSlug": eventSlug})
      if not final_json:
        final_json = json_data
      else:
        final_json['tournament']['events'][0]['entrants']['nodes'] += json_data['tournament']['events'][0]['entrants']['nodes']
      tmp_counter += 10
      tmp_page_counter += 1
    return final_json

if __name__ == "__main__":
    # Check if compute was the given argument
    if len(sys.argv) == 2:
      if sys.argv[1] != 'compute':
        print("Usage: python app.py <tournament_slug> <event_slug> <tournament_tier> or python app.py compute")
        sys.exit(1)
      else:
        compute_ranking()
        sys.exit(0)
    # Check if a slug was provided as a command-line argument
    elif len(sys.argv) != 4:
        print("Usage: python app.py <tournament_slug> <event_slug> <tournament_tier> or python app.py compute")
        sys.exit(1)

    # Get tournament slug and retrieve needed data 
    tournament_slug = sys.argv[1]
    event_slug = sys.argv[2]
    final_data = fetch_tournament(tournament_slug, event_slug)
    
    json_data = {"tournaments": []}

    if os.path.exists('tournament_data.json'):
      with open('tournament_data.json', 'r') as file:
        json_data = json.load(file)

    tmp_data = {
    "name": final_data['tournament']['name'],
    "entrants": final_data['tournament']['events'][0]['entrants']['pageInfo']['total'],
    "tier": sys.argv[3]
    }
    
    json_data['tournaments'].append(tmp_data)
    
    tournament_df_old = None
    head2head_df_old = None
    
    # If excel already created, add new tournament to data
    if os.path.exists('ranking_data.xlsx'):
      tournament_df_old = pd.read_excel('ranking_data.xlsx', sheet_name='Placements')
      head2head_df_old = pd.read_excel('ranking_data.xlsx', sheet_name='Head-Head', index_col=0)
            
    # Create placement and head-head dataframes with new data
    tournament_df = add_tournament(final_data, tournament_df_old)
    head2head_df = add_head2head(final_data, head2head_df_old)
    
    # Write the DataFrames to different sheets in an Excel file
    with pd.ExcelWriter('ranking_data.xlsx', engine='openpyxl', mode='w') as writer:
      tournament_df.to_excel(writer, sheet_name='Placements', index=False)
      head2head_df.to_excel(writer, sheet_name='Head-Head', index=True)
      
    # Save data to a JSON file
    with open('tournament_data.json', 'w') as json_file:
        json.dump(json_data, json_file, indent=4)

    print("Tournament data has been fetched and stored in tournament_data.json")
    
    # Decorate excel: Set width of columns to see tournament name correctly
    workbook = load_workbook('ranking_data.xlsx')
    
    # Decorate sheet 1
    placement_sheet = workbook['Placements']
    column_names = [cell.value for cell in placement_sheet[1]]
    
    # Set width of columns
    max_width = 0    
    for tmp in range(1, placement_sheet.max_column):
      letter = placement_sheet.cell(row=1, column=tmp + 1).column_letter
      length = len(column_names[tmp]) + 4
      placement_sheet.column_dimensions[letter].width = length
      max_width = length if max_width < length + 2 else max_width
    placement_sheet.column_dimensions['A'].width = max_width
      
    # Decorate sheet 2
    head2head_sheet = workbook['Head-Head'] 
    column_names = [cell.value for cell in head2head_sheet[1]]
    
    # Set width of columns
    for tmp in range(1, head2head_sheet.max_column):
      letter = head2head_sheet.cell(row=1, column=tmp + 1).column_letter
      length = len(column_names[tmp]) + 4
      head2head_sheet.column_dimensions[letter].width = length
      max_width = length if max_width < length + 2 else max_width
    head2head_sheet.column_dimensions['A'].width = max_width
    
    # Save the changes
    workbook.save('ranking_data.xlsx')
    workbook.close()
    
    print("Tournament data has been stored in ranking_data.xlsx")



