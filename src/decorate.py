import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def rgb_to_hex(rgb):
    return "{:02X}{:02X}{:02X}".format(*rgb)

def decorate_placements(sheet):
    # Set width of columns
    column_names = [cell.value for cell in sheet[1]]
    max_width = 0    
    for tmp in range(1, sheet.max_column):
        letter = sheet.cell(row=1, column=tmp + 1).column_letter
        length = len(column_names[tmp]) + 4
        sheet.column_dimensions[letter].width = length
        max_width = length if max_width < length + 2 else max_width
    sheet.column_dimensions['A'].width = max_width
    
def decorate_h2h(sheet, multiplier):   
    # Set width of columns
    column_names = [cell.value for cell in sheet[1]]
    max_width = 0    
    for tmp in range(1, sheet.max_column):
        letter = sheet.cell(row=1, column=tmp + 1).column_letter
        length = len(column_names[tmp]) + 4
        sheet.column_dimensions[letter].width = length
        max_width = length if max_width < length + 2 else max_width
    sheet.column_dimensions['A'].width = max_width
    
    # Loop through all rows and columns in the current sheet
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if isinstance(cell.value, int) or isinstance(cell.value, float):
                cell_value = int(round(cell.value))
                hex_color = ''
                if cell.value > 0:
                    value = 255 - min(cell_value * multiplier, 255)
                    hex_color = rgb_to_hex((value, 255, value))
                elif cell.value < 0:
                    value = 255 - min(abs(cell_value) * multiplier, 255)
                    hex_color = rgb_to_hex((255, value, value))
                else:
                    hex_color = rgb_to_hex((255, 255, 255))

                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.fill = fill

def decorate_excel(excel):
    # Decorate excel: Set width of columns to see tournament name correctly
    workbook = load_workbook(excel)
    
    # Decorate placements sheets
    decorate_placements(workbook['Placements'])
    decorate_placements(workbook['Tournament Score'])
    decorate_placements(workbook['Final Ranking'])
      
    # Decorate h2h sheets
    decorate_h2h(workbook['Head-Head'], 25)
    decorate_h2h(workbook['H2H Score'], 2)
          
    # Save the changes
    workbook.save('ranking_data.xlsx')
    print('Excel has been decorated')
    workbook.close()
    