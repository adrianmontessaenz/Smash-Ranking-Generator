"""Ranking computation and Excel decoration utilities.

This module contains the logic to compute player rankings from
tournament data and helpers to write and style the resulting Excel
workbook. The primary entrypoint for application use is
`compute_ranking(data_manager)` which reads tournaments and players
from a `DataManager` instance and writes `data/final_ranking.xlsx`.
"""

from math import sqrt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import json
import os
import sys

from ranking.data_manager import DataManager


def rgb_to_hex(rgb):
    """Convert an (R, G, B) tuple to an uppercase hex string.

    Args:
        rgb (tuple[int, int, int]): RGB values in 0-255 range.

    Returns:
        str: Hex color string like "FFA07A".
    """
    return "{:02X}{:02X}{:02X}".format(*rgb)

def decorate_placements(sheet):
    """Adjust column widths for placement-style sheets.

    The function inspects the first row for column names and sets
    reasonable widths so long tournament/player names are visible.
    """
    # Set width of columns
    column_names = [cell.value for cell in sheet[1]]
    max_width = 0
    for tmp in range(1, sheet.max_column):
        letter = sheet.cell(row=1, column=tmp + 1).column_letter
        length = len(column_names[tmp]) + 4
        sheet.column_dimensions[letter].width = length
        max_width = length if max_width < length + 2 else max_width
    sheet.column_dimensions['A'].width = max_width
    
def decorate_h2h(sheet, multiplier):
    """Adjust column widths and colorize head-to-head cells.

    Numeric cells are colored green for positive values and red for
    negative values. The `multiplier` controls color intensity scaling.
    """
    # Set width of columns
    column_names = [cell.value for cell in sheet[1]]
    max_width = 0
    for tmp in range(1, sheet.max_column):
        letter = sheet.cell(row=1, column=tmp + 1).column_letter
        length = len(column_names[tmp]) + 4
        sheet.column_dimensions[letter].width = length
        max_width = length if max_width < length + 2 else max_width
    sheet.column_dimensions['A'].width = max_width

    # Loop through all rows and columns in the current sheet
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if isinstance(cell.value, int) or isinstance(cell.value, float):
                cell_value = int(round(cell.value))
                hex_color = ''
                if cell.value > 0:
                    value = 255 - min(cell_value * multiplier, 255)
                    hex_color = rgb_to_hex((value, 255, value))
                elif cell.value < 0:
                    value = 255 - min(abs(cell_value) * multiplier, 255)
                    hex_color = rgb_to_hex((255, value, value))
                else:
                    hex_color = rgb_to_hex((255, 255, 255))

                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.fill = fill

def decorate_excel(excel):
    """Open the workbook at `excel`, adjust formatting, and save it.

    The function modifies placement and head-to-head sheets for
    human-friendly viewing and saves the workbook in place.
    """
    # Decorate excel: Set width of columns to see tournament name correctly
    workbook = load_workbook(excel)

    # Decorate placements sheets
    decorate_placements(workbook['Tournament Data'])
    decorate_placements(workbook['Final Ranking'])
    decorate_placements(workbook['Player Data'])

    # Decorate h2h sheets
    decorate_h2h(workbook['H2H Data'], 25)

    # Save the changes
    workbook.save(excel)
    print('Excel has been decorated')
    workbook.close()
    

def get_podium_bonus(placement) -> float:
    """Return a multiplier bonus for podium placements.

    Args:
        placement (int): Tournament placement (1-based).

    Returns:
        float: Bonus multiplier (1.25 for 1st, 1.15 for 2nd, etc.).
    """
    if placement == 1:
        return 1.25
    elif placement == 2:
        return 1.15
    elif placement == 3:
        return 1.08
    else:
        return 1.0

def compute_ranking(data_manager: DataManager):
    """Compute player rankings from tournaments and write an Excel file.

    The function reads tournament, player and ranking information from
    `data_manager`, computes per-player scores using a weighted
    tournament system and head-to-head adjustments, and writes the
    results to `data/final_ranking.xlsx`. It also calls
    `decorate_excel` to prettify the workbook.

    Args:
        data_manager (DataManager): Source of tournaments and players.

    Returns:
        None
    """
    # Get the list of tournaments and players from the data manager
    tournaments = data_manager.get_tournaments()
    tiers = data_manager.get_tournament_tiers()
    players = data_manager.get_players()
    rankings = data_manager.get_rankings()
    tournament_weights = [0.4, 0.3, 0.2, 0.1]  # Weights for the top 4 tournaments
    
    # Compute each tournament value based on its tier and the number of entrants
    all_player_data = {}
    for tournament in tournaments:
        tier_multiplier = tiers.get(tournament.get("tier", "D"), 1.0)  # Default to C tier if not found
        entrants = tournament.get("entrants", 0)
        tournament_value = entrants * 5
        all_tournament_players = {player["gamerTag"] for player in tournament.get("player_data", [])}
        
        # Look for special players in the tournament
        special_bonus = 0
        for player in all_tournament_players:
            if player in players:
                max_ranking = max(rankings[player_ranking] for player_ranking in players[player])
                special_bonus += max_ranking
        
        # Apply decreasing returns to the special bonus
        tournament_value += special_bonus ** 0.6
        tournament_value *= tier_multiplier
        
        # Compute each player's score based on their placement in the tournament
        for player in tournament.get("player_data", []):
            user_id = player["userId"]
            gamer_tag = player["gamerTag"]
            placement = player["placement"]
            
            if user_id not in all_player_data:
                all_player_data[user_id] = {"gamerTag": [gamer_tag], "ranking_score": 0, "head-to-head-score": 0, "tournament_scores": {}}
            elif gamer_tag not in all_player_data[user_id]["gamerTag"]:
                all_player_data[user_id]["gamerTag"].append(gamer_tag)
            all_player_data[user_id]["tournament_scores"][tournament["name"]] = tournament_value * (((tournament["entrants"] - placement + 1) / tournament["entrants"]) ** 1.7)
            
            # Add player's ranking score to the final list
            if gamer_tag in players:
                all_player_data[user_id]["ranking_score"] = max(all_player_data[user_id]["ranking_score"], max(rankings[player_ranking] for player_ranking in players[gamer_tag]))
    

    # Compute final tournament scores for each player based on the top 3 (or 4) tournaments they participated in
    for player_id, player_data in all_player_data.items():
        tournament_scores = list(player_data["tournament_scores"].values())
        tournament_scores.sort(reverse=True)
        top_scores = tournament_scores[:4]  # Get the top 4 scores
        
        # Select only the weights for the needed tournaments
        active_weights = tournament_weights[:len(tournament_scores)]
        weight_sum = sum(active_weights)
        normalized_weights = [weight / weight_sum for weight in active_weights]
        
        weighted_score = sum(score * weight for score, weight in zip(top_scores, normalized_weights))
        all_player_data[player_id]["final_score"] = weighted_score
        
    # Compute head-to-head scores for each player based on their wins on the tournament data
    for tournament in tournaments:
        for player in tournament.get("player_data", []):
            player_id = player["userId"]
            for opponent in player["matches"]["wins"]:
                opponent_id = opponent["userId"]
                
                # Don't count head-to-head score if the opponent was disqualified
                if not opponent["isDisqualified"]:
                    all_player_data[player_id]["head-to-head-score"] += all_player_data[opponent_id]["final_score"] + all_player_data[opponent_id]["ranking_score"]
    
    # Compute final scores for each player based on their final score and head-to-head score
    for player_id, player_data in all_player_data.items():
        all_player_data[player_id]["final_score"] += sqrt(all_player_data[player_id]["head-to-head-score"]) * 0.5
    
    # Remove players who have participated in less than 3 tournaments from the final ranking and sort the remaining players by their final score in descending order
    filtered_players = {
        player_id: data
        for player_id, data in all_player_data.items()
        if len(data["tournament_scores"]) >= 3
    }
    sorted_players = sorted(
        filtered_players.items(),
        key=lambda x: x[1]["final_score"],
        reverse=True
    )
    
    # Create a DataFrame to store the final ranking data
    ranking_data = []
    for rank, (player_id, player_data) in enumerate(sorted_players, start=1):
        ranking_data.append({
            "Rank": rank,
            "GamerTag": ", ".join(player_data["gamerTag"]),
            "Final Score": player_data["final_score"]
        })
        
    ranking_df = pd.DataFrame(ranking_data)
    
    # Gather player data for excel sheet
    player_excel_data = []
    tournament_excel_data = {}
    head_to_head_excel_data = {}
    for player_id, player_data in all_player_data.items():
        player_excel_data.append({
            "UserId": player_id,
            "GamerTag": ", ".join(player_data["gamerTag"])
        })
    
    # Gather tournament and head-to-head data for excel sheet
    for tournament in tournaments:
        tournament_name = tournament["name"]
        for player in tournament.get("player_data", []):
            player_id = player["userId"]
            gamer_tag = " / ".join(all_player_data[player_id]["gamerTag"])            
            if gamer_tag not in tournament_excel_data:
                tournament_excel_data[gamer_tag] = {}
            
            tournament_excel_data[gamer_tag][tournament_name] = player["placement"]
            
            for opponent in player["matches"]["wins"]:
                loser_id = opponent["userId"]
                loser_tag = " / ".join(all_player_data[loser_id]["gamerTag"])
                
                if gamer_tag not in head_to_head_excel_data:
                    head_to_head_excel_data[gamer_tag] = {}
                
                head_to_head_excel_data[gamer_tag][loser_tag] = (
                    head_to_head_excel_data[gamer_tag].get(loser_tag, 0) + 1
                )
                
            
    # Add data to the excel sheets
    player_df = pd.DataFrame(player_excel_data)
    tournament_df = pd.DataFrame.from_dict(tournament_excel_data, orient="index")
    tournament_df.index.name = "Player"
    tournament_df.reset_index(inplace=True)
    tournament_df.fillna(0, inplace=True)
    
    head_to_head_df = pd.DataFrame.from_dict(head_to_head_excel_data, orient="index")
    head_to_head_df.fillna(0, inplace=True)
    all_players = sorted(set(head_to_head_df.index).union(head_to_head_df.columns))
    head_to_head_df = head_to_head_df.reindex(index=all_players,columns=all_players,fill_value=0)
    head_to_head_df.index.name = "Player"
    head_to_head_df.reset_index(inplace=True)
    
    # Save data on excel
    with pd.ExcelWriter('data/final_ranking.xlsx', engine='openpyxl', mode='w') as writer:
        player_df.to_excel(writer, sheet_name='Player Data', index=False)
        tournament_df.to_excel(writer, sheet_name='Tournament Data', index=False)
        head_to_head_df.to_excel(writer, sheet_name='H2H Data', index=False)
        ranking_df.to_excel(writer, sheet_name='Final Ranking', index=False)
    
    decorate_excel('data/final_ranking.xlsx')
    
        
            
                
    
    